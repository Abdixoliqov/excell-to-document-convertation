from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from app.models import FileUpload
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils.datastructures import MultiValueDictKeyError
from openpyxl import load_workbook
from django.http import FileResponse
from django.conf import settings
import os
from django.db.models import Q

@login_required(login_url="auth:login")
def index(request):
    q = request.GET.get("q") if request.GET.get("q") is not None else ""
    print(q)
    fileupload = FileUpload.objects.filter(Q(name__icontains = q))
    print(fileupload)
    context = {"files": fileupload, "q": q}
    return render(request, "index.html", context=context)

@require_POST
def add_file(request):
    # file = FileForm(request.POST, request.FILES)
    # if file.is_valid():  
    #     handle_uploaded_file(request.FILES['file'])  
    #     return redirect("app:index")
    # else:
    #     print("Xato")
    #     return redirect("app:index")
    try:
        file = request.FILES["excel-file"]
    except MultiValueDictKeyError as e:
        messages.info(request, message="File kiriting!")
        return redirect("app:index")

    if file is not None and str(file.name).endswith(".xlsx"):
        FileUpload.objects.create(excel_file = file)
        return redirect("app:index")
    else:
        messages.info(request, message="Excel file kiriting!")
        return redirect("app:index")

@login_required
def delete_file(request):
    file = FileUpload.objects.all()
    media = settings.MEDIA_ROOT
    for f in file:
        os.remove(os.path.join(media, f.excel_file.name))
    file.delete()
    return redirect("app:index")

def download_file(request, pk):
    obj = get_object_or_404(FileUpload, pk=pk)
    path = obj.excel_file.path
    wb_obj = load_workbook(path)
    sheet_obj = wb_obj.active
    dates = str(sheet_obj.cell(row=2, column=4).value)
    dates = dates[:2] + dates[3:5] + dates[-2:]
    count = 1
    # filename = obj.excel_file.name
    test_path = settings.MEDIA_ROOT.joinpath(f"{obj.name}.txt")
    with open(test_path, "w", encoding='utf-8') as f:
        for index in range(6, sheet_obj.max_row):
            if sheet_obj.cell(row=index, column=3).value:
                first_time = sheet_obj.cell(row=index, column=3).value

            if sheet_obj.cell(row=index, column=5).value:
                num = (count < 10) * 3 * "0" or (count < 100) * 2 * "0" or (count < 1000) * "0" or ""
                ids = sheet_obj.cell(row=index, column=5).value
                blok = sheet_obj.cell(row=index, column=6).value
                line = num + str(count)+ "C" + dates + first_time[:2] + first_time[3:5] + first_time[-2:] + blok + str(ids)
                # print(line)
                f.write(str(line))
                f.write('\n')
                count += 1
    # print(test_path)
    # print(obj.name)
    response = FileResponse(open(test_path, 'rb'), as_attachment=True, filename=test_path.name)
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # response['Content-Disposition'] = f'attachment; filename=""'
    # print(response.get('Content-Disposition'))
    # os.remove(os.path.join(settings.MEDIA_ROOT, f"{obj.name}.txt"))
    return response

def one_object_delete(request, pk):
    obj = get_object_or_404(FileUpload, pk=pk)
    media = settings.MEDIA_ROOT
    os.remove(os.path.join(media, obj.excel_file.name))
    obj.delete()
    return redirect("app:index")
